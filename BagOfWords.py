"""
credits to InsightBot for the Bag of Words tutorial.
Source: http://www.insightsbot.com/blog/R8fu5/bag-of-words-algorithm-in-python-introduction

Agrima Bahl for word frequency:
Source: https://medium.com/@agrimabahl/elegant-python-code-reproduction-of-most-common-words-from-a-story-25f5e28e0f8c
"""

import numpy as np
import re
import collections
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import nltk
# nltk.download('stopwords')
from nltk.corpus import stopwords
import pandas as pd
#import FacebookScraper

class BagOfWords:

	""" 
	Get single column from excel workbook
	"""
	def get_excel_column(self,fpath,col):
		wb = load_workbook(fpath)
		ws = wb.active
		column = ws[col]
		return column


	""" 
	Get single column from csv workbook
	"""
	def get_csv_column(self,fpath,col):
		column = pd.read_csv(fpath,usecols=col)
		return column.values.tolist()


	"""
	Extracts the words from a document using regular expressions. 
	Convert all words to lower case and exclude stop words.
	"""
	def extract_words(self,sentence):
		stop_words = stopwords.words('english')
		words = re.sub("[^\w]", " ",  sentence).split() #nltk.word_tokenize(sentence)
		words_cleaned = [w.lower() for w in words if w.lower() not in stop_words]
		return words_cleaned    

	"""
	1. Builds vocabulary by looping through all documents (sentences) 
	2. extracts the words from each sentence
	3. Removes duplicates using the set function 
	4. Returns a sorted list of words.
	"""
	def tokenize_sentences(self,sentences):
		words = []
		for sentence in sentences:
			w = self.extract_words(sentence)
			words.extend(w)
        
		words = sorted(list(set(words)))
		return words

	"""
	Inputs a sentence and words (our vocabulary). 
	Extracts the words from the input sentence using the previously defined function. 
	Creates a vector of zeros with a length of the number of words in our vocabulary.
	For each word in our sentence, loops through vocabulary and increments count by 1 if the word exists
	"""
	def bag_of_words(self,sentence, words):
		sentence_words = self.extract_words(sentence)
		# frequency word count
		bag = np.zeros(len(words))
		for sw in sentence_words:
			for i,word in enumerate(words):
				if word == sw:
					bag[i] += 1
		return np.array(bag)


	def word_frequency(self,sentences):
		words = []
		for sentence in sentences:
			w = self.extract_words(sentence.value)
			words.extend(w)
		words = sorted(list(words))
		wordcount = collections.defaultdict(int)
		for word in words:
			wordcount[word] +=1
		mc = sorted(wordcount.items(), key=lambda k_v: k_v[1], reverse=True)[:20] # Print top 20 words
		# for word, count in mc:
		# 	print(word, ":", count)
		return mc

	def plot_word_frequency(self,mc):
		mc = dict(mc)
		names = list(mc.keys())
		values = list(mc.values())
		plt.bar(range(len(mc)),values,tick_label=names)
		plt.title('Word Frequency')
		plt.xticks(rotation='vertical')
		plt.ylabel('Frequency')
		plt.xlabel('Words')
		plt.savefig('sorted_word_count.png')
		plt.show()

	def run_bag_of_words(self,sentences):
		sentences = ["Hello","my","name","is","Jeff"]
		vocabulary = self.tokenize_sentences(sentences)
		word_count = bag_of_words("Jeff", vocabulary)

# bow = BagOfWords()
# sentences = ["Arrrgh Rogers.","Yes, I've been a customer for ~22 years, although in that time I've divested myself of all your services except internet.","I don't want your TV service. I don't want your phone service. I don't want your cell service.","I just want internet. Please get that through your thick, corporate skull.","Every time you call me to say", "hey, we can increase your internet speed, and lower the price, but you have to take phone or TV", "I get closer to cancelling internet too.","If you can give me faster and cheaper internet, I'm all for it.","If you have to tack on bullshit services I neither want nor need, then I am not interested. Nor do I want to take the TV box and not use it - I know full well that leaves me on the hook for looking after a box, and from experience, I know how awful returning a piece of equipment can be... records often get ", "lost"," resulting in unfortunate charges, not to mention the general grief of ","I have to look after this until it's time to go back so I don't get a crazy TV bill added to my monthly charges","You clowns are MAKING ME SHOP AROUND YOUR COMPETITORS because the only thing I want less than Rogers TV/Phone is weekly sales calls trying to trick me into taking TV/Phone by fiddling with my internet bill."]
# mc = bow.word_frequency(sentences)
# bow.plot_word_frequency(mc)
# bow.run_bag_of_words(sentences)