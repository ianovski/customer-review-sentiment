from BagOfWords import BagOfWords
from bs4 import BeautifulSoup 
import re
import numpy as np
import pandas as pd
pd.options.mode.chained_assignment = None
from sklearn.model_selection import train_test_split 
import nltk.data
from nltk.corpus import stopwords
from openpyxl import load_workbook
import collections
import matplotlib.pyplot as plt

# The methods review_wordlist() and review_sentences() are written by:
# 	Title: GraphicsDrawer source code
# 	Author: Smith, J
# 	Date: 2011
#	Code version: 2.0
#	Availability: http://www.graphicsdrawer.com


class DataManipulation:
	
	
	def get_csv_column(self,fpath,col):
		"""Get single column from csv workbook"""
		column = pd.read_csv(fpath,usecols=col)
		return column.values.tolist()

	def get_excel_column(self,fpath,col):
		""" Get single column from excel workbook"""
		wb = load_workbook(fpath)
		ws = wb.active
		column = ws[col]
		return column

	def split_dataset(self,fpath,colname):
		"""Split dataset into training and testing sets and save csv files"""
		data = pd.read_csv(fpath)
		y = data[colname]
		X = data.drop(colname,axis=1)
		X_train, X_test, y_train, y_test = train_test_split(X,y,test_size=0.2)
		X_train.to_csv(r'tweet_train_features2.csv',index_label="id",header=True)
		X_test.to_csv(r'tweet_test_features2.csv',index_label="id",header=True)
		y_train.to_csv(r'tweet_train_labels2.csv',header=True)
		y_test.to_csv(r'tweet_test_labels2.csv',header=True)

	def review_wordlist(self,review, remove_stopwords=False):
		"""Remove html tags, non-letters and stopwords from text"""
		# 1. Removing html tags
		review_text = BeautifulSoup(review).get_text()
		# 2. Removing non-letter.
		review_text = re.sub("[^a-zA-Z]"," ",review_text)
		# 3. Converting to lower case and splitting
		words = review_text.lower().split()
		# 4. Optionally remove stopwords
		if remove_stopwords:
			stops = set(stopwords.words("english"))     
			words = [w for w in words if not w in stops]
		return(words)


	def extract_words(self,sentence,additional_stop_words=None):
		"""
		Extracts the words from a document using regular expressions. 
		Convert all words to lower case and exclude stop words.
		"""
		stop_words = stopwords.words('english')
		if(additional_stop_words is not None):
			stop_words = stop_words + additional_stop_words
		words = re.sub("[^\w]", " ",  sentence).split() #nltk.word_tokenize(sentence)
		words_cleaned = [w.lower() for w in words if w.lower() not in stop_words]
		return words_cleaned   

	def tokenize_sentences(self,sentences,file_type):
		"""Return tokenized sentences"""
		words = []
		for sentence in sentences:
			if(file_type == 'xlsx'):
				w = self.extract_words(sentence.value)
			if(file_type == 'csv'):
				w = self.extract_words(sentence[0])
			words.append(w)
		return words

	
	def review_sentences(self,review, tokenizer, remove_stopwords=False):
		"""Split a review into sentences"""
		# 1. Using nltk tokenizer
		# nltk.download('popular')
		tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')
		raw_sentences = tokenizer.tokenize(review.strip())
		sentences = []
		# 2. Loop for each sentence
		for raw_sentence in raw_sentences:
			if len(raw_sentence)>0:
				sentences.append(review_wordlist(raw_sentence,remove_stopwords))

		# This returns the list of lists
		return sentences

	def parse_sentences_from_dataset(self,dataset,colname,tokenizer):
		""""Return cleaned sentences from dataset"""
		sentences = []
		for review in dataset[colname]:
			sentences += self.review_sentences(review, tokenizer)
		return sentences
	
	def word_frequency(self,sentences,additional_stop_words=None):
		"""Return sorted list of most frequent words"""
		words = []
		for sentence in sentences:
			if(additional_stop_words is not None):
				w = self.extract_words(sentence[0],additional_stop_words)
			else:
				w = self.extract_words(sentence[0])
			words.extend(w)
		words = sorted(list(words))
		wordcount = collections.defaultdict(int)
		for word in words:
			wordcount[word] +=1
		sorted_wordcount = sorted(wordcount.items(), key=lambda k_v: k_v[1], reverse=True)[:20] # Print top 20 words
		return sorted_wordcount

	def plot_word_frequency(self,sorted_wordcount):
		"""Plot bar chart of most frequent words"""
		sorted_wordcount = dict(sorted_wordcount)
		names = list(sorted_wordcount.keys())
		values = list(sorted_wordcount.values())
		plt.bar(range(len(sorted_wordcount)),values,tick_label=names)
		plt.title('Word Frequency',fontsize=20)
		plt.xticks(rotation='vertical',fontsize=15)
		plt.ylabel('Frequency',fontsize=18)
		plt.xlabel('Words',fontsize=18)
		plt.savefig('figures/frequency_plot.png')
		plt.tight_layout()
		plt.show()

